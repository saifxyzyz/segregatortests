# # import openpyxl module
import time
import openpyxl

import datetime

# for timezone()
import pytz

# using now() to get current time


from openpyxl import load_workbook
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.styles import numbers

import re
import imaplib2
from copy import copy

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import email
from email.header import decode_header
import os
import datetime

Sellers_List = []

Sender_Address = "saadsm247@gmail.com"
Sender_Pass = "ftsqmmguwjiexrwc"
From = ""


def is_duplicate_file(filename):

    # messages = imap.search(None,'(UNSEEN)')
    session = smtplib.SMTP('smtp.gmail.com', 587)
    session.starttls()
    session.login(Sender_Address, Sender_Pass)
    Receiver_Address=From

    #create email
    message = MIMEMultipart()
    message['From'] = Sender_Address
    message['To'] = Receiver_Address
    # message['Subject'] = f'Warning: this PO has been sent multiple times'

    cwd = os.getcwd()
    for root,dirs,files in os.walk(cwd):
        if filename in files:
            message['Subject'] = f'Warning: this PO has been sent multiple times'

            mail_content=f'''Hello,

Warning: This Product Order has been sent more than once.


Regards,
        '''

                    #Attach Message
            message.attach(MIMEText(mail_content,'plain'))
            text = message.as_string()
            session.sendmail(Sender_Address, Receiver_Address, text)
            print()
            session.quit()
            return True
    return False





def Confirm_Change(Toggle):

    session = smtplib.SMTP('smtp.gmail.com', 587)
    session.starttls()
    session.login(Sender_Address, Sender_Pass)
    Receiver_Address=From

    #create email
    message = MIMEMultipart()
    message['From'] = Sender_Address
    message['To'] = Receiver_Address

    if Toggle:

        #Subject for email
        message['Subject'] = 'Successfully Changed File!'


        #Mail Body
        mail_content=f'''Hello!

You have successfully changed the file for product sellers!

Please find the new product seller file attached in this email.


Regards,

        '''
        #attachments
        message.attach(MIMEText(mail_content,'plain'))
        Attach_File_Path = 'Book.xlsx'
        Attach_File = open(Attach_File_Path,'rb')
        payload = MIMEBase('application', 'octate-stream')
        payload.set_payload((Attach_File).read())
        encoders.encode_base64(payload)
        payload.add_header('Content-Disposition', 'attachments', filename=Attach_File_Path)
        message.attach(payload)

        #send
        text = message.as_string()
        session.sendmail(Sender_Address, Receiver_Address, text)

    else:

        #Subject for email
        message['Subject'] = 'Error Occured in the Last Process!'


        #Mail Body
        mail_content=f'''Hello!

No changes were made to files, please make sure the attachments are correct.

Please find the current product seller file attached in this email for your reference.


Regards,

        '''

        #attachments
        message.attach(MIMEText(mail_content,'plain'))
        Attach_File_Path = 'Book.xlsx'
        Attach_File = open(Attach_File_Path,'rb')
        payload = MIMEBase('application', 'octate-stream')
        payload.set_payload((Attach_File).read())
        encoders.encode_base64(payload)
        payload.add_header('Content-Disposition', 'attachments', filename=Attach_File_Path)
        message.attach(payload)

        #send
        text = message.as_string()
        session.sendmail(Sender_Address, Receiver_Address, text)

    session.quit()








def sendersf(type,poNumber,sentAgain=False):
    #x=["abbar.xlsx", "Al Jameel.xlsx", "Al Khair.xlsx", "Al safi.xlsx", "Baider.xlsx", "Dourra.xlsx", "Simat.xlsx"]

#---------------------------------------------------#





    #create email
    message = MIMEMultipart()
    message['From'] = Sender_Address
    message['To'] = From

    print(sentAgain)

    if type == "purchase" and sentAgain == False:
        message['Subject'] = f'Purchase Order for items with PO Number: {poNumber}'
        mail_content=f'''Hello,

We would like to purchase as attached below.


Regards,

        '''
    elif  type == "purchase" and sentAgain == True :
        message['Subject'] = f'This po has been requested again, Purchase Order for items with the PO Number: {poNumber}'
        mail_content=f'''Hello,

We would like to purchase as attached below.


Regards,

        '''
        print("PO requested again")

    elif  type == "avail" and sentAgain == True :
        message['Subject'] = f'This has been requested again, Check Availability for items with the PO Number: {poNumber}'
        mail_content=f'''Hello,


Please have a look at the attachments of this email to provide availability and prices for the same.


Regards,

'''

    else:
        message['Subject'] = f'Check Availability for items with the PO Number: {poNumber}'
        mail_content=f'''Hello,


Please have a look at the attachments of this email to provide availability and prices for the same.


Regards,

'''
    #Attach Message
    message.attach(MIMEText(mail_content,'plain'))



    for sellers_Path in Sellers_List:


        #Add files to message

        Seller_Path_Seperation = sellers_Path.split('/')


        Attach_File_Path = sellers_Path
        Attach_File = open(Attach_File_Path,'rb')
        payload = MIMEBase('application', 'octate-stream')
        payload.set_payload((Attach_File).read())
        encoders.encode_base64(payload)

        if type == "purchase":
            payload.add_header('Content-Disposition', 'attachments', filename = Seller_Path_Seperation[3])
        if type == "avail":
            payload.add_header('Content-Disposition', 'attachments', filename = Seller_Path_Seperation[3])
        if type == "":
            payload.add_header('Content-Disposition', 'attachments', filename = Seller_Path_Seperation[2])
        message.attach(payload)


    #Login

    session = smtplib.SMTP('smtp.gmail.com', 587)
    session.starttls()
    session.login(Sender_Address, Sender_Pass)


    text = message.as_string()
    session.sendmail(Sender_Address, From, text)
    print("sent email")






#---------------------------------------------------#


#     for sel in Sellers_List:
#         Seller_Path_sep = sel.split('/')
#         Seller_Name = Seller_Path_sep[2].split('.')[0]


#         print(f'{Seller_Name}', end=",")


#         #Email Body, Content
#         mail_content=f'''Hello {Seller_Name},

# Please have a look at the attachments of this email and provide availability and prices for the same.


# Regards,

#         '''

#         #Credentials

#         Receiver_Address = From

#         #create email
#         message = MIMEMultipart()
#         message['From'] = Sender_Address
#         message['To'] = Receiver_Address
#         message['Subject'] = f'Checking Availability for items at {Seller_Name}'

#         #attachments
#         message.attach(MIMEText(mail_content,'plain'))
#         Attach_File_Path = sel
#         Attach_File = open(Attach_File_Path,'rb')
#         payload = MIMEBase('application', 'octate-stream')
#         payload.set_payload((Attach_File).read())
#         encoders.encode_base64(payload)
#         payload.add_header('Content-Disposition', 'attachments', filename=Seller_Path_sep[2])
#         message.attach(payload)

#         #send Mail


#         text = message.as_string()
#         session.sendmail(Sender_Address, Receiver_Address, text)
    print()
    session.quit()





def clean(text):
    # clean text for creating a folder
    x = datetime.datetime.now()
    print(x)
    text = text+"Date:"+str(x)
    return "".join(c if c.isalnum() else "_" for c in text)

def bot(path_of_main,subject,Date="*" , type="" , sentAgain=False, path_of_segregation_file="Book.xlsx"):
    #===========================================================================
    # Give the location of the file

    po_Number=subject
    path = path_of_main
    pathSegFile = path_of_segregation_file
    if not os.path.isdir(f"files/{subject}"):
        os.mkdir(f"files/{subject}")
    if not os.path.isdir(f"files/{subject}/avail"):
        if type == "avail":
            os.mkdir(f"files/{subject}/avail")
    if not os.path.isdir(f"files/{subject}/purchase"):
        if type == "purchase":
            os.mkdir(f"files/{subject}/purchase")
    if not os.path.isdir(f"files/{subject}/misc"):
        if type == "":
            os.mkdir(f"files/{subject}/misc")
    #===========================================================================


    #===========================================================================
    file = openpyxl.load_workbook(path)

    sheet = file.active

    row = sheet.max_row
    column = sheet.max_column




    #===========================================================================

    # --------------------------------------------------------------------------------------

    #first list
    findMainBar = []
    mainBar = []





    #===========================================================================

    #Getting the value of barcodes from main sheet

    #===========================================================================

    for i in range(1, column+1):
        cell = sheet.cell(1,i)
        findMainBar.append(cell.value)
    findMainBarDefault = False
    if findMainBar == ['ASIN', 'External ID', 'Title', 'Quantity Outstanding', 'Total Cost'] or findMainBar == ['ASIN', 'External ID', 'Title', 'Quantity Requested', 'Total Cost']:
        findMainBar = ['Material', 'Barcode', 'Description','QTY/PCS', 'boxes', 'PRICE', 'Total']
        findMainBarDefault = True

    if "External ID" in findMainBar:
        index = findMainBar.index("External ID")+1
    if "Barcode" in findMainBar:
        index = findMainBar.index("Barcode")+1
    if "barcode" in findMainBar:
        index = findMainBar.index("barcode")+1

    for i in range(1,row+1 ):
        cell = sheet.cell(i,index)
        mainBar.append(cell.value)


    def delete(sheet):

        # continuously delete row 2 until there
        # is only a single row left over
        # that contains column names

        for i in range(1, sheet.max_row +1):
            # this method removes the row 2
            sheet.delete_rows(1)
        # return to main function
        return



    #------------------------------------------------------------------------------------------------
    # getting the value of barcode from given list
    #===========================================================================

    #second lists:








    #===========================================================================
    # The final path to create multiple files:
    #===========================================================================






    file2 = openpyxl.load_workbook(pathSegFile, data_only = True)


    nameBarDict = {}
    barLst= []
    tempName=""
    quantityIndex = None






    for sheets in file2:
        ws = file2[sheets.title]

        findBar = []
        Bar = []

          # Finding the barcode inside list of first columns
        row2 = ws.max_row
        column2 = ws.max_column

        # Making a list of the first columns
        for i in range(1, column2+1):
            cell = ws.cell(1,i)
            findBar.append(cell.value)

        if "Barcode" in findBar:
            barIndex = findBar.index("Barcode")+1
        if "barcode" in findBar:
            barIndex = findBar.index("barcode")+1
        if "Bar Code" in findBar:
            barIndex = findBar.index("Bar Code")+1

            #Taking the value of barcode from the given sheet
        for i in range(1,row2+1 ):
            cell = ws.cell(i,barIndex)
            Bar.append(cell.value)

        if "Quantity Per Box" in findBar:
            quantityIndex = findBar.index("Quantity Per Box")+1
        if "Case Qty" in findBar:
            quantityIndex = findBar.index("Case Qty")+1
        if "Cost Per Box" in findBar:
            costPerBoxIndex = findBar.index("Cost Per Box") + 1
        if "Cost Price" in findBar:
            costPerBoxIndex = findBar.index("Cost Price") + 1



        for i in range(1,row2+1 ):
            barCell = ws.cell(i+1,barIndex)
            sellerName =  sheets.title
            quantityPerBoxCell = ws.cell(i+1, quantityIndex)
            costPerBoxCell = ws.cell(i+1, costPerBoxIndex)

            if tempName != sellerName:
                barLst = []
            # [[1,2],[2,5]]
            barLst.append([barCell.value, quantityPerBoxCell.value, costPerBoxCell.value])
            nameBarDict[sellerName] = barLst

            tempName = sellerName
    finalPathb4=""
    finalPath = ""
    finalSheet = ""
    miscLst = []
    for i in range(1,row+1 ):
            cell = sheet.cell(i,index)
            miscLst.append(cell.value)
    for L in nameBarDict:
        for i in range(1,row+1 ):
            cell = sheet.cell(i,index)

            if str(cell.value) in str(nameBarDict[L]):

                if type == "avail":

                    finalPath = f"files/{subject}/avail/{L} - {po_Number} - Availability.xlsx"
                elif type == "purchase":
                    finalPath = f"files/{subject}/purchase/{L} - {po_Number}.xlsx"
                else:
                    finalPath = f"files/{subject}/misc/{L}.xlsx"

                wb = openpyxl.Workbook()

                wb.save(finalPath)
                if finalPath not in Sellers_List:
                    Sellers_List.append(finalPath)

                # print(finalPath)
                finalFile = openpyxl.load_workbook(finalPath)
                finalSheet = finalFile.active

                #===========================================================================
                #deleting previous data:
                #===========================================================================

                if __name__ == '__main__':

                    delete(finalSheet)

                    finalFile.save(finalPath)

                #=============================================================================
                # copy paste template
                # ============================================================================
                def copy_sheet(source_sheet, target_sheet):
                                        copy_cells(source_sheet, target_sheet)  # copy all the cel values and styles
                                        copy_sheet_attributes(source_sheet, target_sheet)


                def copy_sheet_attributes(source_sheet, target_sheet):
                    target_sheet.sheet_format = copy(source_sheet.sheet_format)
                    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
                    target_sheet.merged_cells = copy(source_sheet.merged_cells)
                    target_sheet.page_margins = copy(source_sheet.page_margins)
                    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)

                    ##shoaib --copy company logo-- next 10 lines


                    img = Image('Company Logo/IMDN Logo.jpg')

                    p2e = pixels_to_EMU
                    target_sheet.row_dimensions[4].height = 47.5
                    position = XDRPoint2D(p2e(5), p2e(75))
                    size = XDRPositiveSize2D(p2e(904.32) ,p2e(46.08))

                    img.anchor = AbsoluteAnchor(pos=position,ext=size)
                    target_sheet.add_image(img)


                    # set row dimensions
                    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
                    for rn in range(len(source_sheet.row_dimensions)):
                        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])

                    # if source_sheet.sheet_format.defaultColWidth is None:
                        # print('Unable to copy default column wide')
                    else:
                        target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)

                    # set specific column width and hidden property
                    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
                    for key, value in source_sheet.column_dimensions.items():
                        target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)   # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
                        target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)
                        target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width) # set width for every column
                        target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)


                def copy_cells(source_sheet, target_sheet):
                    for (row, col), source_cell in source_sheet._cells.items():
                        target_cell = target_sheet.cell(column=col, row=row)

                        target_cell._value = source_cell._value
                        target_cell.data_type = source_cell.data_type

                        if source_cell.has_style:
                            target_cell.font = copy(source_cell.font)
                            target_cell.border = copy(source_cell.border)
                            target_cell.fill = copy(source_cell.fill)
                            target_cell.number_format = copy(source_cell.number_format)
                            target_cell.protection = copy(source_cell.protection)
                            target_cell.alignment = copy(source_cell.alignment)

                        if source_cell.hyperlink:
                            target_cell._hyperlink = copy(source_cell.hyperlink)

                        if source_cell.comment:
                            target_cell.comment = copy(source_cell.comment)

                if type == "avail":
                    # print("avail template...")





                    wb_source = openpyxl.load_workbook(r"templates/availabilityN.xlsx", data_only=True)
                    source_sheet = wb_source.active

                    copy_sheet(source_sheet, finalSheet)
                    finalSheet.append(findMainBar)

                    finalFile.save(finalPath)
                    # print("avail template done!")
                elif type == "purchase":
                    # print("purchase template...")





                    wb_source = openpyxl.load_workbook(r"templates/purchase.xlsx", data_only=True)
                    source_sheet = wb_source.active

                    copy_sheet(source_sheet, finalSheet)
                    finalSheet.append(findMainBar)
                    finalFile.save(finalPath)
                    # print("avail template done!")
                else:
                    finalSheet.append(findMainBar)
                    finalFile.save(finalPath)


        #===========================================================================
        #===========================================================================
        #adding the given data to file:
        #===========================================================================
        totalForSheet = 0
        for i in range(1,row+1 ):

            cell = sheet.cell(i,index)
            if str(cell.value) in str(nameBarDict[L]):
                if cell.value in miscLst:
                    miscLst.pop(miscLst.index(cell.value))
                tempLst = []
                for j in range(1, column+1):
                    cell2 = sheet.cell(mainBar.index(cell.value)+1,j)

                    if cell2.value != []:
                        tempLst.append(cell2.value)


                # ['6290050509644', 24, 12], ['5000318111424', 72, 8.6], ['8000255115254', 12, 12.1]], 'Al Jameel': [['6291100343904', 24, 9.8], ['6287011312979', 6, 78.5]]
                for i in nameBarDict[L]:

                    if str(i[0]) == str(cell.value):
                        nameBarDictCellValueIndex = nameBarDict[L].index(i)

                if findMainBarDefault == True:

                    noOfBoxes = tempLst[3]/nameBarDict[L][nameBarDictCellValueIndex][1]

                    tempLst.insert(4,noOfBoxes)
                    cost = nameBarDict[L][nameBarDictCellValueIndex][2] * nameBarDict[L][nameBarDictCellValueIndex][1]
                    tempLst.insert(5,cost)
                    tempLst.pop()
                    tempLst.append(cost*noOfBoxes)

                    noOfBoxes = 0
                    cost = 0
                print(tempLst)
                finalSheet.append(tempLst)

                # print(finalPath)
                finalFile.save(finalPath)


        # runLst=[]
        # for i in range(1,row+1 ):

        #     cell = sheet.cell(i,index)
        #     if cell.value in nameBarDict[L]:
        if type == "purchase":

            if finalPath != finalPathb4:
                char_to_num_dict = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K', 12: 'L', 13: 'M', 14: 'N', 15: 'O', 16: 'P', 17: 'Q', 18: 'R', 19: 'S', 20: 'T', 21: 'U', 22: 'V', 23: 'W', 24: 'X', 25: 'Y', 26: 'Z'}
                max_row = finalSheet.max_row
                max_col = finalSheet.max_column
                sumCell = finalSheet.cell(column=max_col, row=max_row+1)
                        # cellBeforeTarget_cell = finalSheet.cell(column=max_col, row=max_row)

                        # print(cellBeforeTarget_cell.value, f"=SUM({char_to_num_dict[max_col]}1:G{max_row-1})" )
                        # if cellBeforeTarget_cell.value != f"=SUM({char_to_num_dict[max_col]}1:G{max_row-1})" :


                sumCell.value = f"=SUM({char_to_num_dict[max_col]}1:G{max_row})"
                target_cell = finalSheet.cell(column=max_col-1, row=max_row+1)
                target_cell.value = f"Sub Total"
                target_cell = finalSheet.cell(column=max_col, row=max_row+2)
                target_cell.value = f"=({char_to_num_dict[max_col]}{max_row+1} *0.15)"
                target_cell = finalSheet.cell(column=max_col-1, row=max_row+2)
                target_cell.value = f"VAT 15%"
                target_cell = finalSheet.cell(column=max_col, row=max_row+3)
                target_cell.value = f"=SUM({char_to_num_dict[max_col]}{max_row+1}, {char_to_num_dict[max_col]}{max_row+2} )"
                target_cell = finalSheet.cell(column=max_col-1, row=max_row+3)
                target_cell.value = f"Total"
                for eachrows in finalSheet[12:finalSheet.max_row]:
                    eachrows[1].number_format = numbers.FORMAT_NUMBER








                finalFile.save(finalPath)

        print("finalSheet:", finalSheet)
        if finalSheet != "":
            maxrow = finalSheet.max_row
            maxcolumn = finalSheet.max_column

            for eachcol in finalSheet[11]:
                eachcol.fill = PatternFill(start_color="D9D9D9",end_color="D9D9D9",fill_type="solid")


            dims = {}
            for row2 in finalSheet.rows:
                for cell2 in row2:
                    if cell2.value:
                        dims[cell2.column_letter] = max((dims.get(cell2.column_letter, 0), len(str(cell2.value))))+1
            for col2, value2 in dims.items():
                finalSheet.column_dimensions[col2].width = value2
            finalSheet.column_dimensions['B'].width = 90


            for row1 in finalSheet[11:maxrow]:
                for cell1 in row1[0:maxcolumn]:
                    cell1.alignment = Alignment(horizontal='center')
                    cell1.border = Border(left=Side(style='thin'),
                                          right=Side(style='thin'),
                                          top=Side(style='thin'),
                                          bottom=Side(style='thin'))
            finalFile.save(finalPath)



        for eachrow in finalSheet:
            for cell in eachrow:
                if cell.value== "VarVendor":
                    if finalPath != finalPathb4:
                        cell.value = L
                        finalFile.save(finalPath)
                if cell.value== "VarPO":
                    if finalPath != finalPathb4:
                        cell.value = po_Number
                        finalFile.save(finalPath)
                if cell.value== "VarDate":
                    if finalPath != finalPathb4:
                        cell.value = Date
                        finalFile.save(finalPath)




        finalPathb4 = finalPath


    print(miscLst)
    if miscLst != []:
        print("misc products found")
        print("making misc file...")
        finalPathMisc = ""
        if type == "purchase":
            finalPathMisc =  f"files/{subject}/purchase/miscellaneous.xlsx"
        elif type == "avail":
            finalPathMisc = f"files/{subject}/avail/miscellaneous.xlsx"
        else:
            finalPathMisc = f"files/{subject}/miscellaneous/miscellaneous.xlsx"
        wbMisc = openpyxl.Workbook()

        wbMisc.save(finalPathMisc)

        finalFileMisc = openpyxl.load_workbook(finalPathMisc)
        finalSheetMisc = finalFileMisc.active
        if __name__ == '__main__':
            delete(finalSheetMisc)
            finalFile.save(finalPathMisc)


        for i in range(1,row+1 ):

            cell = sheet.cell(i,index)

            if str(cell.value) in str(miscLst):

                tempLst = []
                for j in range(1, column+1):
                    cell2 = sheet.cell(mainBar.index(cell.value)+1,j)

                    if cell2.value != []:
                        tempLst.append(cell2.value)
                finalSheetMisc.append(tempLst)

                # print(finalPath)
                finalFileMisc.save(finalPathMisc)

        Sellers_List.append(finalPathMisc)
        # wb_source = openpyxl.load_workbook("templates\calculations.xlsx", data_only=True)
        # source_sheet = wb_source.active

        # copy_sheet(source_sheet, finalSheet)



        # for (row, col), source_cell in source_sheet._cells.items():
        #                 max_row += 1
        #                 target_cell = finalSheet.cell(column=col, row=row)

        #                 target_cell._value = source_cell._value
        #                 target_cell.data_type = source_cell.data_type

    sendersf(type,po_Number,sentAgain)




# bot("input/segregate/myPo.xlsx", "tests")

# import imapclient
# import pyzmail

# i = imapclient.IMAPClient('imap.gmail.com')

# i.login(email, Sender_Pass)
# i.select_folder('INBOX')
# uids = i.search(['all'])
# rawmsgs = i.fetch(uids, ['BODY[]'])
# for i in uids:
#     msg = pyzmail.PyzMessage.factory(rawmsgs[i][b'BODY[]'])

#     print("subject:",msg.get_subject())
#     print("From:",msg.get_addresses('from'))
#     print("To:",msg.get_addresses('to'))
#     print("CC:",msg.get_addresses('cc'))
#     print("Content",msg.text_part.get_payload().decode(msg.text_part.charset))




def imap():
    while True:
        global From
        print("Starting...")
        imap_server = "imap.gmail.com"


        imap = imaplib2.IMAP4_SSL(imap_server)


        imap.login(Sender_Address, Sender_Pass)
        print("logged in")
        imap.select("INBOX")
        while True:
            try:
                print("going idle...")

                imap.idle(120)
                current_time = datetime.datetime.now(pytz.timezone('Asia/Kolkata'))

                # printing current time in india
                print("The current time in india is :", current_time)
                Sellers_List.clear()
                status, messages = imap.search(None,'(UNSEEN)')

                for i in messages[0].split():
                    res, msg = imap.fetch(i, "(RFC822)")
                    for response in msg:
                        if isinstance(response, tuple):

                            msg = email.message_from_bytes(response[1])

                            subject, encoding = decode_header(msg["Subject"])[0]

                            if isinstance(subject, bytes):

                                subject = subject.decode(encoding)

                            From, encoding = decode_header(msg.get("From"))[0]
                            if isinstance(From, bytes):
                                From = From.decode(encoding)
                            print("Subject:", subject)
                            print("From:", From)

                            if msg.is_multipart():

                                for part in msg.walk():

                                    content_type = part.get_content_type()

                                    content_disposition = str(part.get("Content-Disposition"))
                                    try:

                                        body = part.get_payload(decode=True).decode()
                                    except:
                                        pass
                                    if content_type == "text/plain" and "attachment" not in content_disposition:
                                        print(body)
                                        # if "Change_file" or "change_file" or "Segregate_file" or "segregate_file" or "seperate_file" or "Seperate_file" in subject:
                                        #     Confirm_Change(0)

                                    elif "attachment" in content_disposition:
                                        # download attachment
                                        filename = part.get_filename()
                                        if filename:

                                            # if is_duplicate_file(filename):
                                            #     print("\nDuplicate Warning Sent"*5)
                                            #     continue
                                            def extract_action_words(sentence):
                                                pattern = r'action:\s*"([^"]+)"'
                                                match = re.search(pattern, sentence)
                                                if match:
                                                    subject_word = match.group(1)
                                                    return subject_word
                                                else:
                                                    return ""
                                            def extract_type_words(sentence):
                                                pattern = r'type:\s*"([^"]+)"'
                                                match = re.search(pattern, sentence)
                                                if match:
                                                    subject_word = match.group(1)
                                                    return subject_word
                                                else:
                                                    return ""

                                            def extract_purchase_order_words(sentence):
                                                pattern = r'PONo:\s*"([^"]+)"'
                                                match = re.search(pattern, sentence)
                                                if match:
                                                    subject_word = match.group(1)
                                                    return subject_word
                                                else:
                                                    return ""
                                            def extract_date(sentence):
                                                pattern = r'date:\s*"([^"]+)"'
                                                match = re.search(pattern, sentence)
                                                if match:
                                                    subject_word = match.group(1)
                                                    return subject_word
                                                else:
                                                    return ""

                                            extracted_type = extract_type_words(subject)
                                            extracted_action = extract_action_words(subject)
                                            extracted_purchase_order = extract_purchase_order_words(subject)
                                            extracted_date = extract_date(subject)
                                            print(extracted_type,extracted_action,extracted_purchase_order, extracted_date)
                                            if "segregate_file" in extracted_action:
                                                folder_name = extracted_purchase_order

                                                if not os.path.isdir(f"input/{folder_name}"):
                                                    # make a folder for this email (named after the subject)
                                                    os.mkdir(f"input/{folder_name}")

                                                if "avail" in extracted_type:
                                                    if not os.path.isdir(f"input/{folder_name}/avail"):
                                                    # make a avail folder for this email
                                                        os.mkdir(f"input/{folder_name}/avail")
                                                        filepath = os.path.join("input/",f"{folder_name}/","avail/" , filename)
                                                        open(filepath, "wb").write(part.get_payload(decode=True))
                                                        print(filepath)
                                                        bot(filepath,folder_name,extracted_date,"avail")
                                                        print("sent avail to bot with:",filepath,folder_name,"avail")
                                                    else:
                                                        filepath = os.path.join("input/",f"{folder_name}/","avail/" , filename)
                                                        open(filepath, "wb").write(part.get_payload(decode=True))
                                                        print(filepath)
                                                        bot(filepath,folder_name,extracted_date,"avail", True)
                                                        print("sent avail to bot with:",filepath,folder_name,"avail")
                                                elif "purchase" in extracted_type:
                                                    if not os.path.isdir(f"input/{folder_name}/purchase"):
                                                    # make a purchase folder for this email

                                                        os.mkdir(f"input/{folder_name}/purchase")
                                                        filepath = os.path.join("input/",f"{folder_name}/","purchase/" , filename)
                                                        open(filepath, "wb").write(part.get_payload(decode=True))
                                                        print(filepath)
                                                        bot(filepath,folder_name,extracted_date,"purchase")
                                                    else:
                                                        filepath = os.path.join("input/",f"{folder_name}/","purchase/" , filename)
                                                        open(filepath, "wb").write(part.get_payload(decode=True))
                                                        print(filepath)
                                                        bot(filepath,folder_name,extracted_date,"purchase",True)
                                                        print("set sent again to True")

                                                else:
                                                    bot(filepath,folder_name,extracted_date)
                                                    print("sent purchase to bot with:",filepath,folder_name,"purchase")
                                            elif "Change_file" in extracted_action or "change_file" in extracted_action :
                                                if filename == "Book.xlsx":
                                                    filepath =  filename

                                                    if os.path.isfile(filepath):
                                                        os.remove(filepath)
                                                    # download attachment and save it
                                                    open(filepath, "wb").write(part.get_payload(decode=True))
                                                    Confirm_Change(1)
                                                else:
                                                    Confirm_Change(0)





                            else:
                                # extract content type of email
                                content_type = msg.get_content_type()
                                # get the email body
                                body = msg.get_payload(decode=True).decode()
                                if content_type == "text/plain":
                                    # print only text email parts
                                    print(body)
                            # if content_type == "text/html":
                            #     # if it's HTML, create a new HTML file and open it in browser
                            #     folder_name = clean(subject)
                            #     if not os.path.isdir(folder_name):
                            #         # make a folder for this email (named after the subject)
                            #         os.mkdir(folder_name)
                            #     filename = "index.html"
                            #     filepath = os.path.join(folder_name, filename)
                            #     # write the file
                            #     open(filepath, "w").write(body)
                            #     # open in the default browser
                            #     webbrowser.open(filepath)
                            print("="*50)

        # close the connection and logout
            except(imap.abort):
                print("Error has been handled!!!")
                print("closing imap connection")
                print(imap.abort)

                imap.logout()
                break

imap()
